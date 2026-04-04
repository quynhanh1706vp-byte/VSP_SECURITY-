#!/usr/bin/env python3
"""VSP SIEM Excel export — incidents, rules, playbooks, assets, IOCs"""
import argparse, json, sys
from datetime import datetime
from urllib.request import urlopen, Request
import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter

API = "http://127.0.0.1:8921"

# Colors
BG_DARK   = "0a0c10"
BG_SURF   = "1e2128"
BG_HEAD   = "111318"
C_RED     = "ef4444"
C_AMBER   = "f59e0b"
C_GREEN   = "22c55e"
C_BLUE    = "3b82f6"
C_CYAN    = "06b6d4"
C_PURPLE  = "8b5cf6"
C_T1      = "e8eaf0"
C_T2      = "9aa3b8"
C_T3      = "5a6278"
WHITE     = "FFFFFF"

def fetch(path, token):
    try:
        req = Request(f"{API}{path}", headers={"Authorization": f"Bearer {token}"})
        with urlopen(req, timeout=10) as r:
            return json.loads(r.read())
    except Exception as e:
        print(f"⚠ {path}: {e}", file=sys.stderr)
        return {}

def hfill(color): return PatternFill("solid", fgColor=color)
def bold(size=10, color=WHITE): return Font(bold=True, size=size, color=color, name="Calibri")
def norm(size=9,  color=C_T1):  return Font(size=size, color=color, name="Calibri")
def mono(size=8,  color=C_CYAN): return Font(size=size, color=color, name="Courier New")
def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)
def thin_border():
    s = Side(style="thin", color=C_T3)
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, row, cols, bg=BG_HEAD):
    for c, (text, width) in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.font      = bold(9, WHITE)
        cell.fill      = hfill(bg)
        cell.alignment = center()
        cell.border    = thin_border()
        ws.column_dimensions[get_column_letter(c)].width = width

def write_row(ws, row, values, bg=BG_SURF, alt=False):
    bg_use = "111318" if alt else BG_SURF
    for c, (val, font, align) in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=val)
        cell.font      = font
        cell.fill      = hfill(bg_use)
        cell.alignment = align
        cell.border    = thin_border()

def add_title(ws, title, subtitle=""):
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 18
    c1 = ws.cell(row=1, column=1, value=title)
    c1.font = Font(bold=True, size=16, color=WHITE, name="Calibri")
    c1.fill = hfill(BG_DARK)
    c1.alignment = left()
    if subtitle:
        c2 = ws.cell(row=2, column=1, value=subtitle)
        c2.font = Font(size=9, color=C_T3, name="Calibri")
        c2.fill = hfill(BG_DARK)
    return 3  # next data row

def sev_color(sev):
    return {"CRITICAL":C_RED,"HIGH":C_AMBER,"MEDIUM":C_BLUE,"LOW":C_GREEN}.get(
        (sev or "").upper(), C_T2)

def build_excel(token, output):
    print("Fetching data...")
    incidents = fetch("/api/v1/correlation/incidents?limit=200", token)
    rules     = fetch("/api/v1/correlation/rules", token)
    playbooks = fetch("/api/v1/soar/playbooks", token)
    runs      = fetch("/api/v1/soar/runs?limit=100", token)
    sources   = fetch("/api/v1/logs/sources", token)
    iocs      = fetch("/api/v1/ti/iocs?limit=100", token)
    assets    = fetch("/api/v1/assets", token)
    baseline  = fetch("/api/v1/ueba/baseline", token)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = C_RED
    ws.sheet_view.showGridLines = False

    # Title
    ws.row_dimensions[1].height = 40
    c = ws.cell(row=1, column=1,
        value=f"VSP SIEM Report — {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.font = Font(bold=True, size=18, color=WHITE, name="Calibri")
    c.fill = hfill(BG_DARK)
    c.alignment = left()
    ws.merge_cells("A1:H1")

    # KPI cards row
    kpis = [
        ("Incidents",      len(incidents.get("incidents",[])),  C_RED),
        ("Rules",          len(rules.get("rules",[])),          C_BLUE),
        ("Playbooks",      len(playbooks.get("playbooks",[])),  C_PURPLE),
        ("Log Sources",    len(sources.get("sources",[])),      C_GREEN),
        ("IOCs",           len(iocs.get("iocs",[])),            C_CYAN),
        ("Assets",         len(assets.get("assets",[])),        C_AMBER),
        ("Avg Score",      round(baseline.get("avg_score",0)),  C_CYAN),
        ("Pass Rate",
         f"{round(baseline.get('gate_pass_rate',0)*100)}%",     C_GREEN),
    ]
    ws.row_dimensions[2].height = 14
    ws.row_dimensions[3].height = 32
    ws.row_dimensions[4].height = 18
    for i, (lbl, val, color) in enumerate(kpis, 1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = 16
        lc = ws.cell(row=2, column=i, value=lbl)
        lc.font = Font(size=8, color=C_T3, name="Calibri")
        lc.fill = hfill(BG_SURF)
        lc.alignment = center()
        vc = ws.cell(row=3, column=i, value=val)
        vc.font = Font(bold=True, size=18, color=color, name="Calibri")
        vc.fill = hfill(BG_SURF)
        vc.alignment = center()
        bc = ws.cell(row=4, column=i, value="")
        bc.fill = hfill(BG_SURF)

    # Summary table
    r = 6
    ws.cell(row=r, column=1, value="SIEM Coverage Summary").font = bold(11, WHITE)
    ws.cell(row=r, column=1).fill = hfill(BG_DARK)
    r += 1
    summary_rows = [
        ("Correlation Engine", f"{len(rules.get('rules',[]))} rules, "
         f"{sum(1 for x in rules.get('rules',[]) if x.get('enabled'))} enabled"),
        ("Active Incidents",   f"{len(incidents.get('incidents',[]))} total ("
         f"{sum(1 for x in incidents.get('incidents',[]) if (x.get('severity','')+'').upper()=='CRITICAL')} critical)"),
        ("SOAR Playbooks",     f"{len(playbooks.get('playbooks',[]))} configured, "
         f"{sum(1 for x in playbooks.get('playbooks',[]) if x.get('enabled'))} enabled"),
        ("Log Ingestion",      f"{len(sources.get('sources',[]))} sources"),
        ("Threat Intel",       f"{len(iocs.get('iocs',[]))} IOCs loaded"),
        ("Asset Inventory",    f"{len(assets.get('assets',[]))} assets tracked"),
        ("UEBA Baseline",      f"Avg score: {round(baseline.get('avg_score',0))}/100, "
         f"Pass rate: {round(baseline.get('gate_pass_rate',0)*100)}%"),
        ("Generated",          datetime.now().strftime("%Y-%m-%d %H:%M UTC")),
    ]
    for lbl, val in summary_rows:
        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font = bold(9, C_T2); lc.fill = hfill(BG_DARK); lc.alignment = left()
        vc = ws.cell(row=r, column=2, value=val)
        vc.font = norm(9, C_T1); vc.fill = hfill(BG_SURF); vc.alignment = left()
        ws.merge_cells(f"B{r}:H{r}")
        r += 1

    # ── Sheet 2: Incidents ────────────────────────────────────
    ws2 = wb.create_sheet("Incidents")
    ws2.sheet_properties.tabColor = C_RED
    ws2.sheet_view.showGridLines = False
    r = add_title(ws2, "Active Incidents", "Correlation engine detections")
    cols = [("ID",8),("Title",45),("Severity",12),("Status",10),
            ("Rule",25),("Created",18)]
    write_header(ws2, r, cols); r += 1
    for i, inc in enumerate(incidents.get("incidents",[])):
        sev = (inc.get("severity") or "").upper()
        write_row(ws2, r, [
            (str(inc.get("id",""))[:8],         mono(8, C_CYAN), left()),
            (str(inc.get("title","—"))[:80],     norm(9, C_T1),  left()),
            (sev,                                Font(bold=True,size=9,color=sev_color(sev),name="Calibri"), center()),
            (str(inc.get("status","open")),      norm(9, C_T2),  center()),
            (str(inc.get("rule_name","—"))[:30], norm(8, C_T2),  left()),
            (str(inc.get("created_at",""))[:19], mono(8, C_T3),  center()),
        ], alt=i%2==1); r += 1

    # ── Sheet 3: Correlation Rules ────────────────────────────
    ws3 = wb.create_sheet("Correlation Rules")
    ws3.sheet_properties.tabColor = C_BLUE
    ws3.sheet_view.showGridLines = False
    r = add_title(ws3, "Correlation Rules", "Cross-source event correlation")
    cols = [("Name",45),("Severity",12),("Window(min)",12),
            ("Sources",25),("Condition",35),("Enabled",10),("Hits",8)]
    write_header(ws3, r, cols); r += 1
    for i, rule in enumerate(rules.get("rules",[])):
        sev = (rule.get("severity") or "").upper()
        write_row(ws3, r, [
            (str(rule.get("name",""))[:60],                    norm(9,C_T1), left()),
            (sev, Font(bold=True,size=9,color=sev_color(sev),name="Calibri"), center()),
            (rule.get("window_min",0),                         mono(9,C_T2), center()),
            (", ".join(rule.get("sources",[]))[:30],           norm(8,C_T2), left()),
            (str(rule.get("cond",""))[:50],                    mono(8,C_T3), left()),
            ("Yes" if rule.get("enabled") else "No",
             Font(bold=True,size=9,color=C_GREEN if rule.get("enabled") else C_T3,name="Calibri"), center()),
            (rule.get("hits",0),                               mono(9,C_T2), center()),
        ], alt=i%2==1); r += 1

    # ── Sheet 4: SOAR ─────────────────────────────────────────
    ws4 = wb.create_sheet("SOAR Playbooks")
    ws4.sheet_properties.tabColor = C_PURPLE
    ws4.sheet_view.showGridLines = False
    r = add_title(ws4, "SOAR Playbooks", "Automated response workflows")
    cols = [("Name",40),("Trigger",20),("Sev Filter",12),
            ("Enabled",10),("Runs",8),("Success",10),("Rate",8)]
    write_header(ws4, r, cols); r += 1
    for i, pb in enumerate(playbooks.get("playbooks",[])):
        runs_n = pb.get("runs",0)
        succ   = pb.get("success",0)
        rate   = f"{int(succ/max(runs_n,1)*100)}%" if runs_n else "—"
        write_row(ws4, r, [
            (str(pb.get("name",""))[:50],          norm(9,C_T1), left()),
            (str(pb.get("trigger","—")),            mono(8,C_T2), left()),
            (str(pb.get("sev","any")),              norm(8,C_T2), center()),
            ("Yes" if pb.get("enabled") else "No",
             Font(bold=True,size=9,color=C_GREEN if pb.get("enabled") else C_T3,name="Calibri"), center()),
            (runs_n,                                mono(9,C_T2), center()),
            (succ,                                  mono(9,C_T2), center()),
            (rate,  Font(bold=True,size=9,color=C_GREEN if rate=="100%" else C_AMBER,name="Calibri"), center()),
        ], alt=i%2==1); r += 1

    # ── Sheet 5: IOCs ─────────────────────────────────────────
    ws5 = wb.create_sheet("Threat Intel IOCs")
    ws5.sheet_properties.tabColor = C_CYAN
    ws5.sheet_view.showGridLines = False
    r = add_title(ws5, "Threat Intelligence — IOCs", "Indicators of Compromise")
    cols = [("Type",10),("Value",50),("Severity",12),
            ("Feed",18),("Description",40),("Matched",10)]
    write_header(ws5, r, cols); r += 1
    for i, ioc in enumerate(iocs.get("iocs",[])):
        sev = (ioc.get("sev") or "").upper()
        matched = ioc.get("matched", False)
        write_row(ws5, r, [
            (str(ioc.get("type","")).upper(),       mono(8,C_CYAN), center()),
            (str(ioc.get("val",""))[:80],           mono(8,C_T1),  left()),
            (sev, Font(bold=True,size=9,color=sev_color(sev),name="Calibri"), center()),
            (str(ioc.get("feed","—")),              norm(8,C_T2),  center()),
            (str(ioc.get("desc",""))[:60],          norm(8,C_T3),  left()),
            ("✓" if matched else "—",
             Font(bold=True,size=11,color=C_GREEN if matched else C_T3,name="Calibri"), center()),
        ], alt=i%2==1); r += 1

    # ── Sheet 6: Assets ───────────────────────────────────────
    ws6 = wb.create_sheet("Assets")
    ws6.sheet_properties.tabColor = C_AMBER
    ws6.sheet_view.showGridLines = False
    r = add_title(ws6, "Asset Inventory", "CMDB — discovered and manual assets")
    cols = [("Name",30),("Type",12),("IP/Host",20),("Env",10),
            ("Critical",10),("High",10),("Total Findings",14),("Risk Score",12)]
    write_header(ws6, r, cols); r += 1
    for i, a in enumerate(assets.get("assets",[])):
        risk = a.get("risk_score",0)
        write_row(ws6, r, [
            (str(a.get("name",""))[:35],   norm(9,C_T1),  left()),
            (str(a.get("type","host")),     mono(8,C_T2),  center()),
            (str(a.get("ip") or a.get("host",""))[:25], mono(8,C_T3), left()),
            (str(a.get("env","prod")),      norm(8,C_T2),  center()),
            (a.get("critical",0),
             Font(bold=True,size=9,color=C_RED if a.get("critical",0)>0 else C_T3,name="Calibri"), center()),
            (a.get("high",0),
             Font(bold=True,size=9,color=C_AMBER if a.get("high",0)>0 else C_T3,name="Calibri"), center()),
            (a.get("total_findings",0),    norm(8,C_T2),  center()),
            (risk,
             Font(bold=True,size=9,
                  color=C_RED if risk>=70 else C_AMBER if risk>=40 else C_GREEN,
                  name="Calibri"), center()),
        ], alt=i%2==1); r += 1

    # ── Sheet 7: Run History ──────────────────────────────────
    ws7 = wb.create_sheet("SOAR Runs")
    ws7.sheet_properties.tabColor = C_GREEN
    ws7.sheet_view.showGridLines = False
    r = add_title(ws7, "SOAR Playbook Run History", "Execution log")
    cols = [("Run ID",16),("Playbook",35),("Status",12),
            ("Trigger",18),("Started",18),("Finished",18)]
    write_header(ws7, r, cols); r += 1
    for i, run in enumerate(runs.get("runs",[])):
        status = run.get("status","—")
        write_row(ws7, r, [
            (str(run.get("id",""))[:12],     mono(8,C_CYAN), left()),
            (str(run.get("pb","—"))[:45],    norm(9,C_T1),  left()),
            (status,
             Font(bold=True,size=9,
                  color=C_GREEN if status=="success" else C_RED,
                  name="Calibri"), center()),
            (str(run.get("trigger","—")),    norm(8,C_T2),  center()),
            (str(run.get("ts",""))[:19],     mono(8,C_T3),  center()),
            (str(run.get("finished_at","") or "running")[:19], mono(8,C_T3), center()),
        ], alt=i%2==1); r += 1

    # ── Save ──────────────────────────────────────────────────
    wb.save(output)
    import os
    print(f"✓ Excel saved: {output} ({os.path.getsize(output):,} bytes)")
    print(f"  Sheets: {[s.title for s in wb.worksheets]}")

if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--token",  required=True)
    p.add_argument("--output", default="vsp_siem_report.xlsx")
    args = p.parse_args()
    build_excel(args.token, args.output)
